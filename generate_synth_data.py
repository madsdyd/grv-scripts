#!/usr/bin/env python3

"""
This script can generate synthetic data for members, based on some input data.
"""

import pandas as pd
import random
from faker import Faker
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Setup Faker and constants
fake = Faker("da_DK")
Faker.seed(42)

# Settings
num_records = 120
output_path = "members-synthetic.xlsx"

# Postal codes and associated streets for Gladsaxe kommune
postal_codes_distribution = {
    "2800": "Kongens Lyngby",
    "2860": "Søborg",
    "2880": "Bagsværd"
}
distribution_weights = [0.1, 0.35, 0.55]
roads_by_postal_code = {
    "2800": ["Nybro Vænge", "Sandkrogen", "Mogens Alle", "Møllevænget", "Nørgaardsvej", "Stengårds Allé", "Møllestien", "Christoffers Alle", "Gammelmosevej", "Stengårdsvænge", "Amundsensvej"],
    "2860": ["Andersen Nexø Vej", "Bakkedraget", "Fremtidsvej", "Frødings Alle", "Gladsaxevej", "Mørkhøjvej",  "Buddingevej", "Ærtemarken", "Grønnemose Alle", "Gyngemose Parkvej", "Helleskrænten", "Holmevej", "Høje Gladsaxe", "Højvangen", "Juni Alle", "Kildebakkegårds Alle", "Kildebakken", "Kiplings Alle", "Lykkeborgs Alle", "Magledalen", "Maglestien", "Marienborg Alle", "Mørkhøj Parkalle", "Niels Finsens Alle", "Nordtoftevej", "Rugmarken", "Rundgården", "Selma Lagerløfs Alle", "Stengårds Alle", "Søborg Hovedgade", "Vandtårnsvej", "Wergelands Alle"],
    "2880": ["Aldershvilevej","Bagsværddal","Bagsværd Hovedgade","Bindeledet","Bondehavevej","Elmevænget","Frodesvej","Frøstjernevej","Gammelmosevej","Hanehøj","Helmsvej","Krogmosevej","Marsk Stigs Alle","Nydamsvej","Rylevænget","Råvænget","Skovalleen","Skovdiget","Skovkanten","Solbærvænget","Tordisvej","Vandkarsevej","Værebrovej","Østerhegn","Værebrovej", "Vadstrupvej"],
}

# Danish-sounding, "social-liberal/woke" themed names
first_names = ["Aksel", "Astrid", "Freja", "Emil", "Lærke", "Rasmus", "Karen", "Malthe", "Sofie", "Mikkel", "Karla", "Sigrid", "Viggo", "Jens", "Emma", "Gustav", "Katinka", "Matilde", "Martin"]
last_names_affixes = ["Frihed", "Lighed", "Velfærd", "Klima", "Tolerans", "Bro", "Mangfold", "Liberal", "Social", "Åben", "Kreativ", "Forandring", "Udvikling"]

# Helper functions
def random_date(start_year=1970, end_year=2005):
    start_date = datetime(start_year, 1, 1)
    end_date = datetime(end_year, 12, 31)
    delta = end_date - start_date
    random_days = random.randint(0, delta.days)
    return start_date + timedelta(days=random_days)

def generate_danish_phone():
    number = f"{random.randint(10000000, 99999999)}"  # 8-digit Danish number
    if random.random() > 0.7:  # 30% chance to add "+45"
        number = f"+45 {number}"
    return number if random.random() > 0.2 else ""  # 20% chance for an empty field

# Generate synthetic data
data = []
for i in range(1, num_records + 1):
    member_type = random.choice(["Medlem", "Støttemedlem", "Æresmedlem"])
    start_date = random_date(2020, 2024)
    member_no = f"{10000000 + i}"
    first_joined = random_date(2010, 2020)
    
    # Generate name with a "social-liberal/woke" style
    first_name = random.choice(first_names)
    last_name = f"{random.choice(last_names_affixes)}{random.choice(last_names_affixes)}"
    name = f"{first_name} {last_name}"
    
    email = f"{first_name.lower()}.{last_name.lower()}@example.com"
    mobile = generate_danish_phone()
    phone = generate_danish_phone()
    
    # Assign postal code, city, and appropriate street
    postal_code, city = random.choices(
        list(postal_codes_distribution.items()), weights=distribution_weights, k=1
    )[0]
    # Vi begrænser til 100. Øger sandsynligheden for at danne en faktisk adresse.
    street = f"{random.choice(roads_by_postal_code[postal_code])} {random.randint(1, 100)}"
    postal_city = f"{postal_code} {city}"
    
    birthday = random_date(1950, 2000) if random.random() > 0.2 else None
    
    data.append({
        "Medlemstype": member_type,
        "Startdato for aktuel medlemstype": start_date,
        "Medlemsnr.": member_no,
        "Første indmeldelsesdato": first_joined,
        "Navn": name,
        "E-mail": email,
        "Mobil": mobile,
        "Telefon": phone,
        "Vej, husnr. og evt. etage": street,
        "Postnummer og by": postal_city,
        "Fødselsdag": birthday
    })

# Convert to DataFrame
df = pd.DataFrame(data)

# Adjust date columns to proper format without time
for date_col in ["Startdato for aktuel medlemstype", "Første indmeldelsesdato", "Fødselsdag"]:
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

# Save DataFrame to Excel with openpyxl for additional formatting
df.to_excel(output_path, index=False, engine='openpyxl', startrow=1)  # Start headers from row 2

# Load workbook and add header modifications with formatting
wb = load_workbook(output_path)
ws = wb.active

# Insert "Adresse" in merged cells on the first row
ws.merge_cells("I1:J1")
ws["I1"].value = "Adresse"

# Set date columns to proper date format without time
for cell in ws["B2:B" + str(ws.max_row)] + ws["D2:D" + str(ws.max_row)] + ws["K2:K" + str(ws.max_row)]:
    cell[0].number_format = "DD-MM-YYYY"  # Set to date format without time

# Save the final, formatted workbook
wb.save(output_path)
print(f"Synthetic data saved to {output_path}")
