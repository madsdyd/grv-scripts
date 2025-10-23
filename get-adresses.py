#!/usr/bin/env python3

import requests
import json
import sys
import argparse
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time
import traceback

API_BASE = "https://api.dataforsyningen.dk"

INTRO_SHEET_CONTENT = [
    "Dette regneark indeholder en oversigt over vejnavne og antallet af registrerede postkasser pr. vej i hver kommune.",
    "Data er hentet fra DAWA (Danmarks Adressers Web API).",
    "Arket er genereret automatisk af scriptet get-adresses.py.",
    "Hver faneblad repræsenterer en enkelt kommune, sorteret alfabetisk.",
    "",
    "Fejl og mangler kan forekomme – se evt. https://dawa.aws.dk for mere info."
]

def get_kommunekode(kommunenavn: str) -> str:
    url = f"{API_BASE}/kommuner"
    r = requests.get(url, params={"navn": kommunenavn}, timeout=10)
    r.raise_for_status()
    data = r.json()
    if not data:
        raise ValueError(f"Ingen kommune fundet med navnet '{kommunenavn}'.")

    exact = [k for k in data if k.get("navn", "").lower() == kommunenavn.lower()]
    chosen = exact[0] if exact else data[0]
    if not exact and len(data) > 1:
        sys.stderr.write(f"[advarsel] Fandt ikke en entydig match for '{kommunenavn}'. Vælger '{chosen.get('navn')}' (kode {chosen.get('kode')}).")
    return chosen["kode"], chosen.get("navn", kommunenavn)

def stream_adresser(kommunekode: str, per_side: int = 1000):
    base = f"{API_BASE}/adresser"
    params = {"kommunekode": kommunekode, "struktur": "mini"}

    while True:
        sys.stderr.write("Henter side... ")
        sys.stderr.flush()
        try:
            r = requests.get(base, params=params, timeout=30)
            if r.status_code == 400:
                sys.stderr.write("400 fra API – antager at der ikke er flere sider.\n")
                break
            r.raise_for_status()
            items = r.json()
        except requests.HTTPError as e:
            if getattr(e.response, "status_code", None) == 400:
                sys.stderr.write("400 fra API – antager at der ikke er flere sider.\n")
                break
            sys.stderr.write(f"HTTP-fejl: {e}\n")
            break
        except Exception as e:
            sys.stderr.write(f"Fejl: {e}\n")
            break

        if not items:
            sys.stderr.write("Ingen flere data.\n")
            break

        sys.stderr.write(f"{len(items)} adresser hentet.\n")
        sys.stderr.flush()

        for it in items:
            yield it

        if len(items) < per_side:
            sys.stderr.write("Sidste side nået (færre end per_side).\n")
            break

        break
        time.sleep(2)

def retrying_stream_adresser(kommunekode, retries=3, delay=3):
    for attempt in range(retries):
        try:
            yield from stream_adresser(kommunekode)
            return
        except Exception as e:
            if attempt < retries - 1:
                sys.stderr.write(f"Fejl under hentning, prøver igen ({attempt+1}/{retries})...\n")
                time.sleep(delay)
            else:
                raise

def hent_alle_kommuner() -> list[tuple[str, str]]:
    r = requests.get(f"{API_BASE}/kommuner", timeout=10)
    r.raise_for_status()
    data = r.json()
    return sorted([(k["kode"], k["navn"]) for k in data], key=lambda x: x[1])

def generate_excel(outfile, name_pattern=None):
    wb = Workbook()
    intro = wb.active
    intro.title = "Introduktion"
    for idx, line in enumerate(INTRO_SHEET_CONTENT, start=1):
        intro.cell(row=idx, column=1, value=line)

    kommuner = hent_alle_kommuner()
    if name_pattern:
        regex = re.compile(name_pattern, flags=re.IGNORECASE)
        kommuner = [k for k in kommuner if regex.match(k[1])]

    for kode, navn in kommuner:
        sys.stderr.write(f"Behandler {navn} (kode {kode})...\n")
        counts = Counter()
        try:
            for it in retrying_stream_adresser(kode):
                vej = it.get("vejnavn", "")
                counts[vej] += 1
        except Exception:
            traceback.print_exc()
            sys.stderr.write(f"Springer {navn} over pga. fejl.\n")
            continue

        ws = wb.create_sheet(title=navn[:31])
        ws.append(["Vejnavn", "Antal postkasser"])
        for vej in sorted(counts.keys(), key=lambda s: (s is None, s)):
            ws.append([vej, counts[vej]])

        ws.auto_filter.ref = f"A1:B{ws.max_row}"
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 30
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(outfile)
    sys.stderr.write(f"Skrev Excel-fil til: {outfile}\n")

def main():
    parser = argparse.ArgumentParser(description=(
        "Hent adresser fra DAWA for en kommune eller generér samlet Excel-ark for alle."
    ))
    parser.add_argument("kommunenavn", nargs="?", help="Navn på kommune, fx 'Furesø'")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--long", action="store_true", help="Output alle felter som NDJSON (en JSON pr. linje)")
    group.add_argument("--detail", action="store_true", help="Output 'betegnelse' pr. linje")
    parser.add_argument("--outfile", help="Skriv output til fil i stedet for stdout")
    parser.add_argument("--per-page", type=int, default=1000, help="Antal elementer pr. side (default 1000)")
    parser.add_argument("--all-excel", action="store_true", help="Generér Excel med data for alle kommuner")
    parser.add_argument("--match", help="Regulært udtryk for kommunenavne ved --all-excel")

    args = parser.parse_args()

    if args.all_excel:
        if not args.outfile:
            sys.stderr.write("Fejl: --outfile er påkrævet med --all-excel\n")
            sys.exit(1)
        generate_excel(args.outfile, name_pattern=args.match)
        sys.exit(0)

    if not args.kommunenavn:
        sys.stderr.write("Fejl: Du skal angive et kommunenavn, medmindre du bruger --all-excel\n")
        sys.exit(1)

    try:
        kommunekode, resolved_name = get_kommunekode(args.kommunenavn)
    except Exception as e:
        sys.stderr.write(f"Fejl ved opslag af kommune: {e}\n")
        sys.exit(1)

    sys.stderr.write(f"Kommune: {resolved_name} (kode {kommunekode})\n")
    out = open(args.outfile, "w", encoding="utf-8") if args.outfile else sys.stdout

    try:
        if args.long:
            for it in stream_adresser(kommunekode, per_side=args.per_page):
                line = json.dumps(it, ensure_ascii=False)
                out.write(line + "\n")
        elif args.detail:
            for it in stream_adresser(kommunekode, per_page=args.per_page):
                bet = it.get("betegnelse", "")
                out.write(f"{bet}\n")
        else:
            counts = Counter()
            for it in stream_adresser(kommunekode, per_side=args.per_page):
                vej = it.get("vejnavn", "")
                counts[vej] += 1
            for vej in sorted(counts.keys(), key=lambda s: (s is None, s)):
                out.write(f"{vej}\t{counts[vej]}\n")
    finally:
        if args.outfile and out is not sys.stdout:
            out.close()

if __name__ == "__main__":
    main()

